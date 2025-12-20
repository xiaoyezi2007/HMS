import csv
import io
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Any, Dict, Tuple

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlmodel import select
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from app.api.deps import get_current_admin_user
from app.core.config import get_session
from app.core.security import get_password_hash
from app.models.user import UserAccount, UserRole
from app.models.hospital import Doctor, Nurse, Department, Ward, Gender, Payment, Patient
from app.schemas.user import StaffAccountCreate, UserAccountSafe
from app.schemas.hospital import (
    DoctorTitleUpdate,
    NurseHeadUpdate,
    ALLOWED_DOCTOR_TITLES,
    DepartmentCreate,
    WardCreate,
    WARD_TYPE_RULES,
)

router = APIRouter()

MANAGEABLE_ROLES = {UserRole.DOCTOR, UserRole.NURSE, UserRole.PHARMACIST}
DEFAULT_PASSWORD = "123456"
TEMPLATE_DATA_ROWS = 500

IMPORT_TEMPLATE_HEADERS: List[Tuple[str, str]] = [
    ("phone", "手机号 (必填)"),
    ("username", "用户名/昵称 (必填)"),
    ("role", "角色：医生/护士/药师"),
    ("dept_id", "医生所属科室 ID (必填，可在“科室列表”工作表查看)"),
    ("doctor_name", "医生姓名"),
    ("doctor_gender", "医生性别：男/女"),
    ("doctor_title", "医生级别：普通医师/专家医师"),
    ("nurse_name", "护士姓名"),
    ("nurse_gender", "护士性别：男/女")
]

COLUMN_ALIAS_MAP = {
    "phone": "phone",
    "手机号": "phone",
    "username": "username",
    "昵称": "username",
    "role": "role",
    "角色": "role",
    "dept_id": "dept_id",
    "科室id": "dept_id",
    "department": "dept_id",
    "doctor_name": "doctor_name",
    "医生姓名": "doctor_name",
    "doctor_gender": "doctor_gender",
    "医生性别": "doctor_gender",
    "doctor_title": "doctor_title",
    "医生级别": "doctor_title",
    "nurse_name": "nurse_name",
    "护士姓名": "nurse_name",
    "nurse_gender": "nurse_gender",
    "护士性别": "nurse_gender"
}

ROLE_ALIAS_MAP = {
    UserRole.DOCTOR.value: UserRole.DOCTOR,
    "doctor": UserRole.DOCTOR,
    "医生": UserRole.DOCTOR,
    UserRole.NURSE.value: UserRole.NURSE,
    "nurse": UserRole.NURSE,
    "护士": UserRole.NURSE,
    UserRole.PHARMACIST.value: UserRole.PHARMACIST,
    "pharmacist": UserRole.PHARMACIST,
    "药师": UserRole.PHARMACIST
}

GENDER_ALIAS_MAP = {
    Gender.MALE.value: Gender.MALE,
    "男": Gender.MALE,
    "male": Gender.MALE,
    Gender.FEMALE.value: Gender.FEMALE,
    "女": Gender.FEMALE,
    "female": Gender.FEMALE
}

SUPPORTED_IMPORT_SUFFIXES = {".csv", ".xlsx"}


def _normalize_column_name(label: Any) -> str:
    text = str(label or "").strip()
    if not text:
        return ""
    lowered = text.lower()
    return COLUMN_ALIAS_MAP.get(lowered, COLUMN_ALIAS_MAP.get(text, lowered))


def _cell_has_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def _build_row_dicts(headers: List[Any], data_rows: List[Any], data_start_row: int) -> List[Tuple[int, Dict[str, Any]]]:
    normalized_headers = [_normalize_column_name(item) for item in headers]
    rows: List[Tuple[int, Dict[str, Any]]] = []
    for row_index, row in enumerate(data_rows, start=data_start_row):
        cells = list(row) if isinstance(row, (list, tuple)) else list(row)
        if not any(_cell_has_value(cell) for cell in cells):
            continue
        record: Dict[str, Any] = {}
        for idx, header in enumerate(normalized_headers):
            if not header:
                continue
            record[header] = cells[idx] if idx < len(cells) else None
        rows.append((row_index, record))
    return rows


def _parse_csv_bytes(content: bytes) -> List[Tuple[int, Dict[str, Any]]]:
    text = content.decode("utf-8-sig")
    csv_file = io.StringIO(text)
    reader = csv.reader(csv_file)
    try:
        headers = next(reader)
    except StopIteration:
        return []
    data_rows = list(reader)
    return _build_row_dicts(headers, data_rows, 2)


def _parse_excel_bytes(content: bytes) -> List[Tuple[int, Dict[str, Any]]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        return []
    headers = list(rows[0])
    data_rows = [list(row) for row in rows[1:]]
    return _build_row_dicts(headers, data_rows, 2)


async def _read_import_rows(upload_file: UploadFile) -> List[Tuple[int, Dict[str, Any]]]:
    suffix = Path(upload_file.filename or "").suffix.lower()
    if suffix not in SUPPORTED_IMPORT_SUFFIXES:
        raise HTTPException(status_code=400, detail="仅支持 .csv 或 .xlsx 模板文件")
    content = await upload_file.read()
    if not content:
        return []
    if suffix == ".csv":
        return _parse_csv_bytes(content)
    return _parse_excel_bytes(content)


def _safe_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _optional_text(value: Any) -> Optional[str]:
    text = _safe_text(value)
    return text or None


def _coerce_int(value: Any) -> Optional[int]:
    text = _safe_text(value)
    if not text:
        return None
    try:
        if "." in text:
            return int(float(text))
        return int(text)
    except ValueError as exc:  # noqa: WPS440
        raise ValueError("科室 ID 必须为数字") from exc


def _normalize_role_value(value: Any) -> UserRole:
    text = _safe_text(value)
    if not text:
        raise ValueError("角色不能为空")
    lowered = text.lower()
    role = ROLE_ALIAS_MAP.get(lowered) or ROLE_ALIAS_MAP.get(text)
    if not role:
        raise ValueError("角色仅支持 医生/护士/药师")
    return role


def _normalize_gender_value(value: Any) -> Gender:
    text = _safe_text(value)
    if not text:
        raise ValueError("性别不能为空")
    lowered = text.lower()
    gender = GENDER_ALIAS_MAP.get(lowered) or GENDER_ALIAS_MAP.get(text)
    if not gender:
        raise ValueError("性别仅支持 男 或 女")
    return gender


def _normalize_doctor_title_text(value: Any) -> Optional[str]:
    text = _safe_text(value)
    if not text:
        return None
    return "专家医师" if text == "主治医师" else text


def _row_to_payload(row: Dict[str, Any]) -> StaffAccountCreate:
    phone = _safe_text(row.get("phone"))
    if not phone:
        raise ValueError("手机号不能为空")
    username = _safe_text(row.get("username"))
    if not username:
        raise ValueError("用户名不能为空")
    role = _normalize_role_value(row.get("role"))
    dept_id = _coerce_int(row.get("dept_id"))
    doctor_gender = None
    nurse_gender = None
    if _safe_text(row.get("doctor_gender")):
        doctor_gender = _normalize_gender_value(row.get("doctor_gender"))
    if _safe_text(row.get("nurse_gender")):
        nurse_gender = _normalize_gender_value(row.get("nurse_gender"))
    doctor_title = _normalize_doctor_title_text(row.get("doctor_title"))
    return StaffAccountCreate(
        phone=phone,
        username=username,
        role=role,
        dept_id=dept_id,
        doctor_name=_optional_text(row.get("doctor_name")),
        doctor_gender=doctor_gender,
        doctor_title=doctor_title,
        nurse_name=_optional_text(row.get("nurse_name")),
        nurse_gender=nurse_gender
    )


def _build_template_workbook(departments: Optional[List[Department]] = None) -> io.BytesIO:
    dept_list = departments or []
    sample_dept_id = dept_list[0].dept_id if dept_list else 1

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "账号导入"
    sheet.append([header for header, _ in IMPORT_TEMPLATE_HEADERS])
    sheet.append([
        "13800000001",
        "张三-医生",
        "医生",
        sample_dept_id,
        "张三",
        "男",
        "专家医师",
        "",
        ""
    ])
    sheet.append([
        "13900000002",
        "李四-护士",
        "护士",
        "",
        "",
        "",
        "",
        "李四",
        "女"
    ])
    sheet.append([
        "13700000003",
        "赵六-药师",
        "药师",
        "",
        "",
        "",
        "",
        "",
        ""
    ])
    sheet.freeze_panes = "A2"

    notes = workbook.create_sheet("须知")
    notes.append(["请勿修改账号导入表的表头，按列填写信息。"])
    notes.append(["手机号、用户名、角色为必填项。医生需填写医生姓名/性别/级别，并在‘医生所属科室 ID’列中填写对应数字，可在‘科室列表’查看。"])
    notes.append(["模板内提供‘科室列表’工作表，可查阅 ID 与名称对应关系。"])
    notes.append(["‘角色’、‘性别’、‘医生级别’列已提供下拉选项，直接选择即可。"])
    notes.append(["允许上传 .xlsx 或 .csv 文件，建议使用此模板直接编辑。"])

    dept_sheet = workbook.create_sheet("科室列表")
    if dept_list:
        dept_sheet.append(["科室ID", "科室名称"])
        for dept in dept_list:
            dept_sheet.append([dept.dept_id, dept.dept_name])
    else:
        dept_sheet.append(["暂无科室数据，请先在系统内创建科室，再填写导入表。"])

    def add_list_validation(formula: str, column_letter: str) -> None:
        dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=True)
        target_range = f"{column_letter}2:{column_letter}{1 + TEMPLATE_DATA_ROWS}"
        sheet.add_data_validation(dv)
        dv.add(target_range)

    role_formula = f'"{UserRole.DOCTOR.value},{UserRole.NURSE.value},{UserRole.PHARMACIST.value}"'
    gender_formula = f'"{Gender.MALE.value},{Gender.FEMALE.value}"'
    doctor_title_formula = f'"{",".join(ALLOWED_DOCTOR_TITLES)}"'

    add_list_validation(role_formula, "C")
    add_list_validation(gender_formula, "F")
    add_list_validation(gender_formula, "I")
    add_list_validation(doctor_title_formula, "G")

    if dept_list:
        dept_range_end = 1 + len(dept_list)
        dept_formula = f"'科室列表'!$A$2:$A${dept_range_end}"
        add_list_validation(dept_formula, "D")
    else:
        notes.append(["暂未配置科室列表，下拉选择不可用，请手动填写科室 ID。"])

    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def _to_safe(account: UserAccount) -> UserAccountSafe:
    return UserAccountSafe(
        phone=account.phone,
        username=account.username,
        role=account.role,
        status=account.status,
        register_time=account.register_time
    )


async def _create_staff_account_record(payload: StaffAccountCreate, session: AsyncSession) -> UserAccount:
    if payload.role not in MANAGEABLE_ROLES:
        raise HTTPException(status_code=400, detail="仅能创建医生/护士/药师账号")

    if payload.role == UserRole.DOCTOR:
        if not payload.dept_id:
            raise HTTPException(status_code=400, detail="医生账号必须指定所属科室")
        if not payload.doctor_name:
            raise HTTPException(status_code=400, detail="请提供医生姓名")
        if not payload.doctor_gender:
            raise HTTPException(status_code=400, detail="请提供医生性别")
        if not payload.doctor_title:
            raise HTTPException(status_code=400, detail="请提供医生级别")

        doctor_title = "专家医师" if payload.doctor_title == "主治医师" else payload.doctor_title
        if doctor_title not in ALLOWED_DOCTOR_TITLES:
            raise HTTPException(status_code=400, detail="医生级别仅支持专家医师或普通医师")

        dept = (
            await session.execute(select(Department).where(Department.dept_id == payload.dept_id))
        ).scalars().first()
        if not dept:
            raise HTTPException(status_code=404, detail="指定的科室不存在")
    else:
        doctor_title = None

    if payload.role == UserRole.NURSE:
        if not payload.nurse_name:
            raise HTTPException(status_code=400, detail="请提供护士姓名")
        if not payload.nurse_gender:
            raise HTTPException(status_code=400, detail="请提供护士性别")

    phone_exists = (
        await session.execute(select(UserAccount).where(UserAccount.phone == payload.phone))
    ).scalars().first()
    if phone_exists:
        raise HTTPException(status_code=400, detail="该手机号已被注册")

    username_exists = (
        await session.execute(select(UserAccount).where(UserAccount.username == payload.username))
    ).scalars().first()
    if username_exists:
        raise HTTPException(status_code=400, detail="该用户名已被占用")

    account = UserAccount(
        phone=payload.phone,
        username=payload.username,
        role=payload.role,
        password_hash=get_password_hash(DEFAULT_PASSWORD),
        status="启用"
    )

    session.add(account)
    await session.commit()
    await session.refresh(account)

    if payload.role == UserRole.DOCTOR:
        doctor = (
            await session.execute(select(Doctor).where(Doctor.phone == payload.phone))
        ).scalars().first()
        gender_value = payload.doctor_gender if isinstance(payload.doctor_gender, Gender) else Gender(payload.doctor_gender)
        if doctor:
            doctor.name = payload.doctor_name
            doctor.gender = gender_value
            doctor.title = doctor_title or ALLOWED_DOCTOR_TITLES[0]
            doctor.dept_id = payload.dept_id
            session.add(doctor)
        else:
            doctor = Doctor(
                name=payload.doctor_name,
                gender=gender_value,
                title=doctor_title or ALLOWED_DOCTOR_TITLES[0],
                phone=payload.phone,
                dept_id=payload.dept_id
            )
            session.add(doctor)
        await session.commit()
        await session.refresh(doctor)

    if payload.role == UserRole.NURSE:
        nurse = (
            await session.execute(select(Nurse).where(Nurse.phone == payload.phone))
        ).scalars().first()
        gender_value = payload.nurse_gender if isinstance(payload.nurse_gender, Gender) else Gender(payload.nurse_gender)
        if nurse:
            nurse.name = payload.nurse_name
            nurse.gender = gender_value
            session.add(nurse)
        else:
            nurse = Nurse(
                name=payload.nurse_name,
                gender=gender_value,
                phone=payload.phone,
                is_head_nurse=False
            )
            session.add(nurse)
        await session.commit()
        await session.refresh(nurse)

    return account


@router.get("/admin/accounts", response_model=List[UserAccountSafe])
async def list_staff_accounts(
    role: Optional[UserRole] = None,
    _: UserAccount = Depends(get_current_admin_user),
    session: AsyncSession = Depends(get_session)
):
    role_filter = role
    if role_filter and role_filter not in MANAGEABLE_ROLES:
        raise HTTPException(status_code=400, detail="只能查看医生/护士/药师账号")

    stmt = select(UserAccount).where(UserAccount.role.in_(list(MANAGEABLE_ROLES)))
    if role_filter:
        stmt = stmt.where(UserAccount.role == role_filter)

    result = await session.execute(stmt)
    accounts = result.scalars().all()
    return [_to_safe(item) for item in accounts]


@router.post("/admin/accounts", response_model=UserAccountSafe, status_code=status.HTTP_201_CREATED)
async def create_staff_account(
    payload: StaffAccountCreate,
    _: UserAccount = Depends(get_current_admin_user),
    session: AsyncSession = Depends(get_session)
):
    account = await _create_staff_account_record(payload, session)
    return _to_safe(account)


@router.delete("/admin/accounts/{phone}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_staff_account(
    phone: str,
    _: UserAccount = Depends(get_current_admin_user),
    session: AsyncSession = Depends(get_session)
):
    stmt = select(UserAccount).where(UserAccount.phone == phone)
    account = (await session.execute(stmt)).scalars().first()
    if not account or account.role not in MANAGEABLE_ROLES:
        raise HTTPException(status_code=404, detail="未找到可删除的医生/护士/药师账号")

    account.status = "禁用"
    session.add(account)
    await session.commit()


@router.get("/admin/accounts/template")
async def download_staff_template(
    _: UserAccount = Depends(get_current_admin_user),
    session: AsyncSession = Depends(get_session)
):
    departments = (await session.execute(select(Department))).scalars().all()
    stream = _build_template_workbook(departments)
    filename = f"staff-account-template-{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    headers = {
        "Content-Disposition": f"attachment; filename={filename}"
    }
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )


@router.post("/admin/accounts/import")
async def import_staff_accounts(
    file: UploadFile = File(...),
    _: UserAccount = Depends(get_current_admin_user),
    session: AsyncSession = Depends(get_session)
):
    rows = await _read_import_rows(file)
    if not rows:
        raise HTTPException(status_code=400, detail="导入文件为空或缺少数据")

    success_items: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []

    for row_number, row_data in rows:
        try:
            payload = _row_to_payload(row_data)
        except ValueError as exc:
            errors.append({"row_number": row_number, "message": str(exc)})
            continue

        try:
            account = await _create_staff_account_record(payload, session)
            success_items.append({
                "row_number": row_number,
                "phone": account.phone,
                "username": account.username,
                "role": account.role.value
            })
        except HTTPException as exc:
            errors.append({"row_number": row_number, "message": str(exc.detail)})
        except Exception as exc:  # noqa: W0703
            errors.append({"row_number": row_number, "message": str(exc)})

    return {
        "total_rows": len(rows),
        "success_count": len(success_items),
        "success_items": success_items,
        "errors": errors
    }


@router.get("/admin/doctors")
async def list_doctors(
    _: UserAccount = Depends(get_current_admin_user),
    session: AsyncSession = Depends(get_session)
):
    # 只返回启用状态的医生账号，并关联科室名称
    stmt = (
        select(Doctor, Department.dept_name)
        .join(UserAccount, Doctor.phone == UserAccount.phone)
        .outerjoin(Department, Doctor.dept_id == Department.dept_id)
        .where(UserAccount.status == "启用")
    )
    result = await session.execute(stmt)
    doctors_with_dept = []
    for doctor, dept_name in result.all():
        doctor_dict = {
            "doctor_id": doctor.doctor_id,
            "name": doctor.name,
            "gender": doctor.gender,
            "title": doctor.title,
            "phone": doctor.phone,
            "dept_id": doctor.dept_id,
            "dept_name": dept_name
        }
        doctors_with_dept.append(doctor_dict)
    return doctors_with_dept


@router.patch("/admin/doctors/{doctor_id}/title", response_model=Doctor)
async def update_doctor_title(
    doctor_id: int,
    payload: DoctorTitleUpdate,
    _: UserAccount = Depends(get_current_admin_user),
    session: AsyncSession = Depends(get_session)
):
    # payload.title 已在 schema validator 中做了兼容与归一化（主治医师 -> 专家医师）
    if payload.title not in ALLOWED_DOCTOR_TITLES:
        raise HTTPException(status_code=400, detail="医生级别仅支持专家医师或普通医师")

    doctor = (await session.execute(select(Doctor).where(Doctor.doctor_id == doctor_id))).scalars().first()
    if not doctor:
        raise HTTPException(status_code=404, detail="医生不存在")

    doctor.title = payload.title
    session.add(doctor)
    await session.commit()
    await session.refresh(doctor)
    return doctor


@router.get("/admin/nurses", response_model=List[Nurse])
async def list_nurses(
    _: UserAccount = Depends(get_current_admin_user),
    session: AsyncSession = Depends(get_session)
):
    # 只返回启用状态的护士账号
    stmt = (
        select(Nurse)
        .join(UserAccount, Nurse.phone == UserAccount.phone)
        .where(UserAccount.status == "启用")
    )
    result = await session.execute(stmt)
    return result.scalars().all()


@router.patch("/admin/nurses/{nurse_id}/head", response_model=Nurse)
async def update_nurse_head_status(
    nurse_id: int,
    payload: NurseHeadUpdate,
    _: UserAccount = Depends(get_current_admin_user),
    session: AsyncSession = Depends(get_session)
):
    nurse = (await session.execute(select(Nurse).where(Nurse.nurse_id == nurse_id))).scalars().first()
    if not nurse:
        raise HTTPException(status_code=404, detail="护士不存在")

    nurse.is_head_nurse = payload.is_head_nurse
    session.add(nurse)
    await session.commit()
    await session.refresh(nurse)
    return nurse


@router.get("/admin/revenue")
async def get_revenue_summary(
    _: UserAccount = Depends(get_current_admin_user),
    session: AsyncSession = Depends(get_session)
):
    paid_filter = Payment.status == "已缴费"

    total_stmt = select(
        func.coalesce(func.sum(Payment.amount), 0).label("total_amount"),
        func.count(Payment.payment_id).label("paid_count")
    ).where(paid_filter)
    total_row = (await session.execute(total_stmt)).first()
    total_amount = float(total_row.total_amount) if total_row and total_row.total_amount is not None else 0.0
    paid_count = total_row.paid_count if total_row else 0

    type_stmt = (
        select(
            Payment.type,
            func.coalesce(func.sum(Payment.amount), 0).label("amount"),
            func.count(Payment.payment_id).label("count")
        )
        .where(paid_filter)
        .group_by(Payment.type)
    )
    type_rows = await session.execute(type_stmt)
    by_type = []
    for row in type_rows:
        entry_type = row.type.value if hasattr(row.type, "value") else row.type
        by_type.append({
            "type": entry_type,
            "amount": float(row.amount or 0.0),
            "count": row.count
        })

    record_stmt = (
        select(Payment, Patient.name, Patient.phone)
        .join(Patient, Payment.patient_id == Patient.patient_id)
        .where(paid_filter)
        .order_by(Payment.time.desc())
        .limit(200)
    )
    record_rows = await session.execute(record_stmt)
    records = []
    for payment, patient_name, patient_phone in record_rows.all():
        records.append({
            "payment_id": payment.payment_id,
            "type": payment.type.value if hasattr(payment.type, "value") else payment.type,
            "amount": payment.amount,
            "time": payment.time,
            "patient_id": payment.patient_id,
            "patient_name": patient_name,
            "patient_phone": patient_phone,
            "pres_id": payment.pres_id,
            "exam_id": payment.exam_id,
            "hosp_id": payment.hosp_id
        })

    return {
        "total_amount": total_amount,
        "paid_count": paid_count,
        "by_type": by_type,
        "records": records
    }


@router.post("/admin/departments", response_model=Department, status_code=status.HTTP_201_CREATED)
async def create_department(
    payload: DepartmentCreate,
    _: UserAccount = Depends(get_current_admin_user),
    session: AsyncSession = Depends(get_session)
):
    existing = (await session.execute(select(Department).where(Department.dept_name == payload.dept_name))).scalars().first()
    if existing:
        raise HTTPException(status_code=400, detail="科室名称已存在")

    dept = Department(dept_name=payload.dept_name, telephone=payload.telephone)
    session.add(dept)
    await session.commit()
    await session.refresh(dept)
    return dept


@router.get("/admin/wards")
async def list_wards(
    _: UserAccount = Depends(get_current_admin_user),
    session: AsyncSession = Depends(get_session)
):
    stmt = select(Ward, Department.dept_name).join(Department, Ward.dept_id == Department.dept_id)
    rows = await session.execute(stmt)
    result = []
    for ward, dept_name in rows.all():
        result.append({
            "ward_id": ward.ward_id,
            "bed_count": ward.bed_count,
            "type": ward.type,
            "dept_id": ward.dept_id,
            "dept_name": dept_name
        })
    return result


@router.post("/admin/wards", status_code=status.HTTP_201_CREATED)
async def create_ward(
    payload: WardCreate,
    _: UserAccount = Depends(get_current_admin_user),
    session: AsyncSession = Depends(get_session)
):
    dept = await session.get(Department, payload.dept_id)
    if not dept:
        raise HTTPException(status_code=404, detail="科室不存在")
    existing = await session.get(Ward, payload.ward_id)
    if existing:
        raise HTTPException(status_code=400, detail="房间号已存在")
    ward_type = payload.type.strip()
    expected_beds = WARD_TYPE_RULES.get(ward_type)
    if expected_beds is None:
        raise HTTPException(
            status_code=400,
            detail=f"病房类型仅支持：{', '.join(WARD_TYPE_RULES.keys())}"
        )

    ward = Ward(ward_id=payload.ward_id, dept_id=payload.dept_id, bed_count=expected_beds, type=ward_type)
    session.add(ward)
    await session.commit()
    await session.refresh(ward)
    return {
        "ward_id": ward.ward_id,
        "bed_count": expected_beds,
        "type": ward.type,
        "dept_id": ward.dept_id,
        "dept_name": dept.dept_name
    }